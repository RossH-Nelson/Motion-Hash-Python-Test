import os
import cv2
import numpy as np
from PIL import Image
import imagehash
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import time
import random

# Function to apply transformations (crops and rotations)
def apply_transformations(img):
    height, width = img.shape[:2]

    mild_crop1 = img[int(0.05 * height):int(0.95 * height), int(0.05 * width):int(0.95 * width)]
    mild_crop2 = img[int(0.1 * height):int(0.9 * height), int(0.1 * width):int(0.9 * width)]
    heavy_crop1 = img[int(0.2 * height):int(0.8 * height), int(0.2 * width):int(0.8 * width)]
    heavy_crop2 = img[int(0.25 * height):int(0.75 * height), int(0.25 * width):int(0.75 * width)]
    random_crop = random_crop_image(img)

    mild_rotation1 = rotate_image(img, 10)
    mild_rotation2 = rotate_image(img, 15)
    heavy_rotation1 = rotate_image(img, 45)
    heavy_rotation2 = rotate_image(img, 90)

    return [(mild_crop1, "Mild Crop 1"), (mild_crop2, "Mild Crop 2"),
            (heavy_crop1, "Heavy Crop 1"), (heavy_crop2, "Heavy Crop 2"),
            (random_crop, "Random Crop"),
            (mild_rotation1, "Mild Rotation 1"), (mild_rotation2, "Mild Rotation 2"),
            (heavy_rotation1, "Heavy Rotation 1"), (heavy_rotation2, "Heavy Rotation 2")]

# Helper function to rotate an image by a specified angle
def rotate_image(img, angle):
    height, width = img.shape[:2]
    matrix = cv2.getRotationMatrix2D((width / 2, height / 2), angle, 1)
    return cv2.warpAffine(img, matrix, (width, height))

# Helper function for random cropping
def random_crop_image(img):
    height, width = img.shape[:2]
    top = random.randint(0, int(0.2 * height))
    bottom = random.randint(int(0.8 * height), height)
    left = random.randint(0, int(0.2 * width))
    right = random.randint(int(0.8 * width), width)
    return img[top:bottom, left:right]

# Refined pHash similarity function with cropping based on the smaller dimensions
def refined_phash_similarity(img1, img2):
    """
    Crops the original image to match the dimensions of the comparison image if needed,
    then compares the pHash similarity.
    """
    if img1.shape != img2.shape:
        # Calculate the center crop of the original image to match comparison image dimensions
        height1, width1 = img1.shape[:2]
        height2, width2 = img2.shape[:2]
        start_x = (width1 - width2) // 2
        start_y = (height1 - height2) // 2
        cropped_img1 = img1[start_y:start_y + height2, start_x:start_x + width2]
    else:
        cropped_img1 = img1  # No need to crop if dimensions are identical

    # Debugging information to confirm dimensions
    print(f"Comparing images of size: {cropped_img1.shape} and {img2.shape}")
    
    # Convert OpenCV images to PIL format for hashing
    pil_img1 = Image.fromarray(cv2.cvtColor(cropped_img1, cv2.COLOR_BGR2RGB))
    pil_img2 = Image.fromarray(cv2.cvtColor(img2, cv2.COLOR_BGR2RGB))
    
    # Compute pHash for both resized images
    hash1 = imagehash.phash(pil_img1)
    hash2 = imagehash.phash(pil_img2)
    
    # Calculate similarity based on pHash (0 = identical, higher values = less similar)
    phash_similarity = 1 - (hash1 - hash2) / len(hash1.hash) ** 2
    return phash_similarity * 100  # Return as percentage similarity

# ORB similarity function that uses keypoint matching
def orb_similarity(img1, img2):
    orb = cv2.ORB_create()
    kp1, des1 = orb.detectAndCompute(img1, None)
    kp2, des2 = orb.detectAndCompute(img2, None)

    if des1 is None or des2 is None:
        return 0

    bf = cv2.BFMatcher(cv2.NORM_HAMMING, crossCheck=True)
    matches = bf.match(des1, des2)

    if len(matches) == 0:
        return 0

    good_matches = [m for m in matches if m.distance < 42]
    similarity_percentage = len(good_matches) / min(len(kp1), len(kp2)) * 100
    return min(similarity_percentage, 100)

# Function to color cells based on percentage
def color_cell_based_on_percentage(cell, percentage):
    color = None
    if 0 <= percentage < 10:
        color = '00FF00'
    elif 10 <= percentage < 20:
        color = '66FF00'
    elif 20 <= percentage < 30:
        color = 'CCFF00'
    elif 30 <= percentage < 40:
        color = 'FFFF00'
    elif 40 <= percentage < 50:
        color = 'FFCC00'
    elif 50 <= percentage < 60:
        color = 'FF9900'
    elif 60 <= percentage < 70:
        color = 'FF6600'
    elif 70 <= percentage < 80:
        color = 'FF3300'
    elif 80 <= percentage < 90:
        color = 'FF0000'
    elif 90 <= percentage <= 100:
        color = '990000'
    
    if color:
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.fill = fill

# Function to process images and generate results
def process_images(image_folder, random_image_folder, output_xlsx):
    image_files = [f for f in os.listdir(image_folder) if f.endswith(('.png', '.jpg', '.jpeg'))]
    random_image_files = [f for f in os.listdir(random_image_folder) if f.endswith(('.png', '.jpg', '.jpeg'))]
    random_image_files = random_image_files[:5]

    wb = Workbook()
    ws = wb.active
    ws.title = "Image Similarities"

    header = ["Original Image", 
              "Mild Crop 1 pHash %", "Mild Crop 1 ORB %", 
              "Mild Crop 2 pHash %", "Mild Crop 2 ORB %", 
              "Heavy Crop 1 pHash %", "Heavy Crop 1 ORB %", 
              "Heavy Crop 2 pHash %", "Heavy Crop 2 ORB %", 
              "Random Crop pHash %", "Random Crop ORB %",
              "Mild Rotation 1 pHash %", "Mild Rotation 1 ORB %", 
              "Mild Rotation 2 pHash %", "Mild Rotation 2 ORB %", 
              "Heavy Rotation 1 pHash %", "Heavy Rotation 1 ORB %", 
              "Heavy Rotation 2 pHash %", "Heavy Rotation 2 ORB %", 
              "Random Image 1 pHash %", "Random Image 1 ORB %", 
              "Random Image 2 pHash %", "Random Image 2 ORB %", 
              "Random Image 3 pHash %", "Random Image 3 ORB %",
              "Random Image 4 pHash %", "Random Image 4 ORB %",
              "Random Image 5 pHash %", "Random Image 5 ORB %"]

    ws.append(header)

    start_time = time.time()
    
    for idx, image_file in enumerate(image_files):
        print(f"Processing image {idx + 1} of {len(image_files)}: {image_file}")
        img_path = os.path.join(image_folder, image_file)
        img = cv2.imread(img_path)
        
        if img is None:
            print(f"Failed to load image: {image_file}")
            continue
        
        transformations = apply_transformations(img)
        row = [image_file]
        
        for transformed_img, transform_name in transformations:
            print(f"  Applying transformation: {transform_name}")
            phash_sim = refined_phash_similarity(img, transformed_img)
            orb_sim = orb_similarity(img, transformed_img)
            row.extend([round(phash_sim, 2), round(orb_sim, 2)])

        for random_image_file in random_image_files:
            random_img_path = os.path.join(random_image_folder, random_image_file)
            random_img = cv2.imread(random_img_path)
            
            if random_img is None:
                print(f"  Failed to load random image: {random_image_file}")
                continue
            
            print(f"  Comparing with random image: {random_image_file}")
            phash_random_sim = refined_phash_similarity(img, random_img)
            orb_random_sim = orb_similarity(img, random_img)
            row.extend([round(phash_random_sim, 2), round(orb_random_sim, 2)])
        
        ws.append(row)
        
        elapsed_time = time.time() - start_time
        print(f"Completed processing for {image_file}. Time elapsed: {elapsed_time:.2f} seconds.\n")

    for row in ws.iter_rows(min_row=2, min_col=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            try:
                value = float(cell.value)
                color_cell_based_on_percentage(cell, value)
            except ValueError:
                pass
    
    wb.save(output_xlsx)

if __name__ == "__main__":
    # Define the folder paths and output file
    image_folder = '/Users/rosshartigan/Nelson Development/Motion Ads/Data-Analysis/Flyers'  # Folder with original 1000 images
    random_image_folder = '/Users/rosshartigan/Nelson Development/Motion Ads/Data-Analysis/Random'  # Folder with random images
    output_xlsx = '/Users/rosshartigan/Nelson Development/Motion Ads/Data-Analysis/CSV/results.xlsx'  # Path to save the Excel file
    
    # Process images
    process_images(image_folder, random_image_folder, output_xlsx)