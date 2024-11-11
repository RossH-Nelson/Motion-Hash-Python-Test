import os
import random
from PIL import Image
import imagehash
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import numpy as np

# Paths
image_folder = '/Users/rosshartigan/Nelson Development/Motion Ads/Data-Analysis/Flyers'
control_folder = '/Users/rosshartigan/Nelson Development/Motion Ads/Data-Analysis/Random'
output_xlsx = '/Users/rosshartigan/Nelson Development/Motion Ads/Data-Analysis/CSV/results.xlsx'
log_folder = '/Users/rosshartigan/Nelson Development/Motion Ads/Data-Analysis/CSV/Logs'

# Global standardized size
STANDARDIZED_SIZE = (720, 720)

# Ensure the log folder exists
os.makedirs(log_folder, exist_ok=True)

# Function to calculate pHash
def calculate_phash(image):
    return imagehash.phash(image)

# Function to calculate hash similarity percentage
def calculate_hash_similarity(hash1, hash2):
    hamming_distance = hash1 - hash2
    hash_length = len(hash1.hash) ** 2
    similarity_percentage = (1 - (hamming_distance / hash_length)) * 100
    return similarity_percentage

# Function to calculate black pixel percentage
def calculate_black_pixel_percentage(image):
    grayscale_image = image.convert("L")
    image_array = np.array(grayscale_image)
    black_pixel_count = np.sum(image_array == 0)
    total_pixels = image_array.size
    black_percentage = (black_pixel_count / total_pixels) * 100
    return round(black_percentage, 2)

# Function to resize and crop image to standardized size
def resize_and_crop(image_path, size=(720, 720)):
    image = Image.open(image_path)
    original_width, original_height = image.size
    target_width, target_height = size

    scale = max(target_width / original_width, target_height / original_height)
    new_width = int(original_width * scale)
    new_height = int(original_height * scale)

    image_resized = image.resize((new_width, new_height), Image.LANCZOS)

    left = (new_width - target_width) // 2
    top = (new_height - target_height) // 2
    right = left + target_width
    bottom = top + target_height

    image_cropped = image_resized.crop((left, top, right, bottom))

    return image_cropped

# Function to apply random crop
def random_crop(image):
    width, height = image.size
    left = random.randint(0, width // 4)
    top = random.randint(0, height // 4)
    right = random.randint(3 * width // 4, width)
    bottom = random.randint(3 * height // 4, height)
    return image.crop((left, top, right, bottom))

# Function to apply random rotation
def random_rotation(image):
    angle = random.randint(1, 359)
    return image.rotate(angle, fillcolor="black")

# Function to apply transformations
def apply_transformations(original_image):
    width, height = original_image.size
    transformations = {
        "Mild Crop 1": original_image.crop((int(0.05 * width), int(0.05 * height), int(0.95 * width), int(0.95 * height))),
        "Mild Crop 2": original_image.crop((int(0.1 * width), int(0.1 * height), int(0.9 * width), int(0.9 * height))),
        "Heavy Crop 1": original_image.crop((int(0.2 * width), int(0.2 * height), int(0.8 * width), int(0.8 * height))),
        "Heavy Crop 2": original_image.crop((int(0.25 * width), int(0.25 * height), int(0.75 * width), int(0.75 * height))),
        "Mild Rotation 1": original_image.rotate(10, fillcolor="black"),
        "Mild Rotation 2": original_image.rotate(15, fillcolor="black"),
        "Heavy Rotation 1": original_image.rotate(45, fillcolor="black"),
        "Heavy Rotation 2": original_image.rotate(90, fillcolor="black"),
        "Random Crop": random_crop(original_image),
        "Random Rotation": random_rotation(original_image),
    }
    return transformations

# Function to color cells based on percentage
def color_cell_based_on_percentage(cell, percentage):
    color = None
    if 0 <= percentage < 10:
        color = '00FF00'
    elif 10 <= percentage < 20:
        color = '66FF00'
    elif 20 <= percentage < 30:
        color = 'CCFF00'
    elif 30 <= percentage < 40:
        color = 'FFFF00'
    elif 40 <= percentage < 50:
        color = 'FFCC00'
    elif 50 <= percentage < 60:
        color = 'FF9900'
    elif 60 <= percentage < 70:
        color = 'FF6600'
    elif 70 <= percentage < 80:
        color = 'FF3300'
    elif 80 <= percentage < 90:
        color = 'FF0000'
    elif 90 <= percentage <= 100:
        color = '990000'
    
    if color:
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.fill = fill

# Function to process images and store data
def process_images(image_folder, control_folder, sample_size, output_xlsx, log_folder):
    files = os.listdir(image_folder)[:sample_size]
    control_files = [os.path.join(control_folder, f) for f in os.listdir(control_folder) 
                     if f.lower().endswith(('.png', '.jpg', '.jpeg'))]
    control_images = [Image.open(f) for f in control_files]
    control_phashes = [calculate_phash(img) for img in control_images]

    results = []
    error_log = []

    for idx, file in enumerate(files, start=1):
        print(f"Processing image {idx} of {len(files)}: {file}")
        file_path = os.path.join(image_folder, file)
        try:
            original_image = Image.open(file_path)
            original_phash = calculate_phash(original_image)
            standardized_image = resize_and_crop(file_path, STANDARDIZED_SIZE)
            standardized_phash = calculate_phash(standardized_image)

            # Compare Standardized pHash to Original pHash
            similarity_standardized_to_original = calculate_hash_similarity(original_phash, standardized_phash)

            transformations = apply_transformations(original_image)
            row = {
                "Image Name": file,
                "Original Size": f"{original_image.size[0]}x{original_image.size[1]}",
                "Standardized % Similarity to Original": round(similarity_standardized_to_original, 2),
            }

            for control_idx, control_phash in enumerate(control_phashes, start=1):
                row[f"Control {control_idx} % Similarity to Original"] = round(calculate_hash_similarity(original_phash, control_phash), 2)
                row[f"Control {control_idx} % Similarity to Standardized"] = round(calculate_hash_similarity(standardized_phash, control_phash), 2)

            for name, transformed_image in transformations.items():
                transformed_phash = calculate_phash(transformed_image)
                row[f"{name} pHash % Similarity to Original"] = round(calculate_hash_similarity(original_phash, transformed_phash), 2)
                row[f"{name} pHash % Similarity to Standardized"] = round(calculate_hash_similarity(standardized_phash, transformed_phash), 2)
                row[f"{name} % Black Pixels"] = calculate_black_pixel_percentage(transformed_image)

            results.append(row)

        except Exception as e:
            print(f"Error processing {file}: {e}")
            error_log.append(f"Error processing {file}: {e}")

    write_to_excel(results, output_xlsx)

    if error_log:
        error_log_path = os.path.join(log_folder, "error_log.txt")
        with open(error_log_path, "w") as log_file:
            log_file.write("\n".join(error_log))
        print(f"Error log saved to {error_log_path}")

# Function to write results to Excel
def write_to_excel(data, output_path):
    workbook = Workbook()
    sheet = workbook.active

    headers = ["Image Name", "Original Size", "Standardized % Similarity to Original"] + \
              [f"Control {i} % Similarity to Original" for i in range(1, 6)] + \
              [f"Control {i} % Similarity to Standardized" for i in range(1, 6)] + \
              [f"{name} pHash % Similarity to Original" for name in [
                  "Mild Crop 1", "Mild Crop 2", "Heavy Crop 1", "Heavy Crop 2",
                  "Mild Rotation 1", "Mild Rotation 2", "Heavy Rotation 1", "Heavy Rotation 2",
                  "Random Crop", "Random Rotation"
              ]] + \
              [f"{name} pHash % Similarity to Standardized" for name in [
                  "Mild Crop 1", "Mild Crop 2", "Heavy Crop 1", "Heavy Crop 2",
                  "Mild Rotation 1", "Mild Rotation 2", "Heavy Rotation 1", "Heavy Rotation 2",
                  "Random Crop", "Random Rotation"
              ]] + \
              [f"{name} % Black Pixels" for name in [
                  "Mild Crop 1", "Mild Crop 2", "Heavy Crop 1", "Heavy Crop 2",
                  "Mild Rotation 1", "Mild Rotation 2", "Heavy Rotation 1", "Heavy Rotation 2",
                  "Random Crop", "Random Rotation"
              ]]
    sheet.append(headers)

    for row in data:
        sheet.append([row.get(header, "") for header in headers])

    for col_idx in range(3, len(headers) + 1):
        for cell in sheet.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2, max_row=sheet.max_row):
            for c in cell:
                try:
                    value = float(c.value)
                    color_cell_based_on_percentage(c, value)
                except ValueError:
                    pass

    workbook.save(output_path)
    print(f"Results saved to {output_path}")

# Main execution
if __name__ == "__main__":
    SAMPLE_SIZE = 1000
    process_images(image_folder, control_folder, SAMPLE_SIZE, output_xlsx, log_folder)